from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


class statsEditor():
     
    
    def __init__(self, master):
        
        
        self.wb = Workbook()
        self.wb = load_workbook('data/volley_stats.xlsx')

        self.ws = self.wb.active
        
        self.weekList = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        self.weekNumber = 0
        
        self.selectedPlayer = "None"

        master.title('Volleyball Statistics Input')
        self.titleLabel = Label(master, text="Volleyball Statistics Input", padx=10, pady=10)
        
        self.weekFrame = LabelFrame(master, text="Week Selection", padx=10, pady=10)
        self.weekLabel = Label(self.weekFrame, text="Week: 1 of "+ str(len(self.weekList)), padx=20, pady=10)
        self.prevWeekButton = Button(self.weekFrame, text= "<<", command=lambda: self.prevWeek(0, self.weekList), padx=10, pady=10, anchor=W) #
        self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(2, self.weekList), padx=10, pady=10, anchor=W) 
    
        #Initializing Player Selection Buttons
        self.playerSelection = LabelFrame(master, text="Player Selection", padx=10, pady=10)
        self.brandonChan = Button(self.playerSelection, text="Chan", command=lambda: self.playerSelect("brandonChan"), height=4, width=15)
        self.callumAshton = Button(self.playerSelection, text="Callum", command=lambda: self.playerSelect("callumAshton"), height=4, width=15)
        self.danielPark = Button(self.playerSelection, text="Daniel", command=lambda: self.playerSelect("danielPark"), height=4, width=15)
        self.deirdreTruong = Button(self.playerSelection, text="Deirdre", command=lambda: self.playerSelect("deirdreTruong"), height=4, width=15)
        self.edwardKang = Button(self.playerSelection, text="Edward", command=lambda: self.playerSelect("edwardKang"), height=4, width=15)
        self.kevinMa = Button(self.playerSelection, text="Kema", command=lambda: self.playerSelect("kevinMa"), height=4, width=15)
        self.kevinTang = Button(self.playerSelection, text="Ktang", command=lambda: self.playerSelect("kevinTang"), height=4, width=15)
        self.lachlanDenham = Button(self.playerSelection, text="Lachlan", command=lambda: self.playerSelect("lachlanDenham"), height=4, width=15)
        self.mimiChen = Button(self.playerSelection, text="Mimi", command=lambda: self.playerSelect("mimiChen"), height=4, width=15)
        self.willOuyang = Button(self.playerSelection, text="Will", command=lambda: self.playerSelect("willOuyang"), height=4, width=15)
        
        #buttonColor = brandonChan.cget("background")
       
        #Initializing Statisitic Add and Subtract Buttons
        self.serveErrors = LabelFrame(master, text="Serve Errors: ", padx=10, pady=10)
        self.serveErrorsAdd = Button(self.serveErrors, text="+", padx=20, pady=15)
        self.serveErrorsRemove = Button(self.serveErrors, text="-", padx=20, pady=15)
  
        self.serveSuccess = LabelFrame(master, text="Serve Successes: ", padx=10, pady=10)
        self.serveSuccessAdd = Button(self.serveSuccess, text="+", padx=20, pady=15)
        self.serveSuccessRemove = Button(self.serveSuccess, text="-", padx=20, pady=15)
  
        self.receiveErrors = LabelFrame(master, text="Receive Errors: ", padx=10, pady=10)
        self.receiveErrorsAdd = Button(self.receiveErrors, text="+", padx=20, pady=15)
        self.receiveErrorsRemove = Button(self.receiveErrors, text="-", padx=20, pady=15)
     
        self.receiveSuccess = LabelFrame(master, text="Receive Successes: ", padx=10, pady=10)
        self.receiveSuccessAdd = Button(self.receiveSuccess, text="+", padx=20, pady=15)
        self.receiveSuccessRemove = Button(self.receiveSuccess, text="-", padx=20, pady=15)
      
        self.spikeErrors = LabelFrame(master, text="Spike Errors: ", padx=10, pady=10)
        self.spikeErrorsAdd = Button(self.spikeErrors, text="+", padx=20, pady=15)
        self.spikeErrorsRemove = Button(self.spikeErrors, text="-", padx=20, pady=15)
     
        self.spikeSuccess = LabelFrame(master, text="Spike Successes: ", padx=10, pady=10)
        self.spikeSuccessAdd = Button(self.spikeSuccess, text="+", padx=20, pady=15)
        self.spikeSuccessRemove = Button(self.spikeSuccess, text="-", padx=20, pady=15)
     
        self.blockErrors = LabelFrame(master, text="Block Errors: ", padx=10, pady=10)
        self.blockErrorsAdd = Button(self.blockErrors, text="+", padx=20, pady=15)
        self.blockErrorsRemove = Button(self.blockErrors, text="-", padx=20, pady=15)
    
        self.blockSuccess = LabelFrame(master, text="Block Successes: ", padx=10, pady=10)
        self.blockSuccessAdd = Button(self.blockSuccess, text="+", padx=20, pady=15)
        self.blockSuccessRemove = Button(self.blockSuccess, text="-", padx=20, pady=15)
     
        self.Faults = LabelFrame(master, text="self.Faults: ", padx=10, pady=10)
        self.FaultsAdd = Button(self.Faults, text="+", padx=20, pady=15)
        self.FaultsRemove = Button(self.Faults, text="-", padx=20, pady=15)
      
        #Initialising Statistics table
        self.statisticsmaster = LabelFrame(master, text="Player Statistics", padx=10, pady=10)
        self.serveErrorsLabel = Label(self.statisticsmaster, text="Serve Errors: ", padx=10, pady=10)
        self.serveSuccessLabel = Label(self.statisticsmaster, text="Serve Successes: ", padx=10, pady=10)
        self.serveRateLabel = Label(self.statisticsmaster, text="Serve Rate: " + "TO FILL" + "%", padx=10, pady=10)
        self.receiveErrorsLabel = Label(self.statisticsmaster, text="Receive Errors: ", padx=10, pady=10)
        self.receiveSuccessLabel = Label(self.statisticsmaster, text="Receive Successes: ", padx=10, pady=10)
        self.receiveRateLabel = Label(self.statisticsmaster, text="Receive Rate: " + "TO FILL" + "%", padx=10, pady=10)
        self.spikeErrorsLabel = Label(self.statisticsmaster, text="Spike Errors: ", padx=10, pady=10)
        self.spikeSuccessLabel = Label(self.statisticsmaster, text="Spike Successes: ", padx=10, pady=10)
        self.spikeRateLabel = Label(self.statisticsmaster, text="Spike Rate: " + "TO FILL" + "%", padx=10, pady=10)
        self.blockErrorsLabel = Label(self.statisticsmaster, text="Block Errors: ", padx=10, pady=10)
        self.blockSuccessLabel = Label(self.statisticsmaster, text="Block Successes: ", padx=10, pady=10)
        self.blockRateLabel = Label(self.statisticsmaster, text="Block Rate: " + "TO FILL" + "%", padx=10, pady=10)
        self.FaultsLabel = Label(self.statisticsmaster, text="self.Faults: ", padx=10, pady=10)
        
        self.exitButton = Button(master, text= "Exit", command=master.destroy, padx=20, pady=10)
        
        ##Attaching all initial state GUI components to grid
        
        #Initial elements
        self.titleLabel.grid(row=0, column=0, padx=(20,0), pady=(0, 20))
        
        self.weekFrame.grid(row=0, column=2, columnspan=3)
        self.prevWeekButton.grid(row=0, column=0)
        self.weekLabel.grid(row=0, column=1)
        self.nextWeekButton.grid(row=0, column=2)
     
        #Player Selection buttons
        self.playerSelection.grid(row=1, column=0, columnspan=5, padx=10, pady=5)
        self.brandonChan.grid(row=0, column=0)
        self.callumAshton.grid(row=0, column=1)
        self.danielPark.grid(row=0, column=2)
        self.deirdreTruong.grid(row=0, column=3)
        self.edwardKang.grid(row=0, column=4)
        self.kevinMa.grid(row=1, column=0)
        self.kevinTang.grid(row=1, column=1)
        self.lachlanDenham.grid(row=1, column=2)
        self.mimiChen.grid(row=1, column=3)
        self.willOuyang.grid(row=1, column=4)
      
        #buttons and frames for changing data
        self.serveErrors.grid(row=2, column=0, padx=10, pady=5)
        self.serveErrorsAdd.grid(row=0, column=0)
        self.serveErrorsRemove.grid(row=0, column=1)
      
        self.serveSuccess.grid(row=2, column=1, padx=10, pady=5)
        self.serveSuccessAdd.grid(row=0, column=0)
        self.serveSuccessRemove.grid(row=0, column=1)
      
        self.receiveErrors.grid(row=2, column=2, padx=10, pady=5)
        self.receiveErrorsAdd.grid(row=0, column=0)
        self.receiveErrorsRemove.grid(row=0, column=1)
      
        self.receiveSuccess.grid(row=2, column=3, padx=10, pady=5)
        self.receiveSuccessAdd.grid(row=0, column=0)
        self.receiveSuccessRemove.grid(row=0, column=1)
   
        self.spikeErrors.grid(row=3, column=0, padx=10, pady=5)
        self.spikeErrorsAdd.grid(row=0, column=0)
        self.spikeErrorsRemove.grid(row=0, column=1)
       
        self.spikeSuccess.grid(row=3, column=1, padx=10, pady=5)
        self.spikeSuccessAdd.grid(row=0, column=0)
        self.spikeSuccessRemove.grid(row=0, column=1)
      
        self.blockErrors.grid(row=3, column=2, padx=10, pady=5)
        self.blockErrorsAdd.grid(row=0, column=0)
        self.blockErrorsRemove.grid(row=0, column=1)
       
        self.blockSuccess.grid(row=3, column=3, padx=10, pady=5)
        self.blockSuccessAdd.grid(row=0, column=0)
        self.blockSuccessRemove.grid(row=0, column=1)
      
        self.Faults.grid(row=4, column=0, padx=10, pady=5)
        self.FaultsAdd.grid(row=0, column=0)
        self.FaultsRemove.grid(row=0, column=1)
      
        #Statistics frame and labels
        self.statisticsmaster.grid(row=4, column=1, columnspan=3, padx=5)
        self.serveErrorsLabel.grid(row=0, column=0, padx=10,)
        self.serveSuccessLabel.grid(row=0, column=1, padx=10)
        self.serveRateLabel.grid(row=0, column=2, padx=10)
        self.receiveErrorsLabel.grid(row=1, column=0, padx=10)
        self.receiveSuccessLabel.grid(row=1, column=1, padx=10)
        self.receiveRateLabel.grid(row=1, column=2, padx=10)
        self.spikeErrorsLabel.grid(row=2, column=0, padx=10)
        self.spikeSuccessLabel.grid(row=2, column=1, padx=10)
        self.spikeRateLabel.grid(row=2, column=2, padx=10)
        self.blockErrorsLabel.grid(row=3, column=0, padx=10)
        self.blockSuccessLabel.grid(row=3, column=1, padx=10)
        self.blockRateLabel.grid(row=3, column=2, padx=10)
        self.FaultsLabel.grid(row=4, column=0, padx=10)
        self.exitButton.grid(row=99, column=4)
    
    #Function for iterating week forwards once
    def nextWeek(self, weekNumber, weekList): 
      
        self.weekLabel.grid_forget()
        self.weekLabel = Label(self.weekFrame, text="Week: " + str(weekList[weekNumber-1]) + " of " + str(len(weekList)), padx=20, pady=10, anchor=W)
        self.prevWeekButton = Button(self.weekFrame, text= "<<", command=lambda: self.prevWeek(weekList[weekNumber-2], weekList), padx=10, pady=10, anchor=W)
            
        if(weekNumber >= len(weekList)):
            self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(1, weekList), padx=10, pady=10, anchor=W)
            self.prevWeekButton = Button(self.weekFrame, text= "<<", command=lambda: self.prevWeek(len(weekList)-1, weekList), padx=10, pady=10, anchor=W)
        else:
            self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(weekList[weekNumber], weekList), padx=10, pady=10, anchor=W)
        
        self.weekLabel.grid(row=0, column=1)
        self.prevWeekButton.grid(row=0, column=0)
        self.nextWeekButton.grid(row=0, column=2)
    
    #Function for iterating week backwards once
    def prevWeek(self, weekNumber, weekList):
      
        self.weekLabel.grid_forget()
        self.weekLabel = Label(self.weekFrame, text="Week: " + str(weekList[weekNumber-1]) + " of " + str(len(weekList)), padx=20, pady=10, anchor=W)
        self.prevWeekButton = Button(self.weekFrame, text= "<<", command=lambda: self.prevWeek(weekList[weekNumber-2], weekList), padx=10, pady=10, anchor=W)
        
        if(weekNumber >= len(weekList)):
            self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(1, weekList), padx=10, pady=10, anchor=W)
        else:
            self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(weekList[weekNumber], weekList), padx=10, pady=10, anchor=W)
        
        self.weekLabel.grid(row=0, column=1)
        self.prevWeekButton.grid(row=0, column=0)
        self.nextWeekButton.grid(row=0, column=2)
        return
    
    #Function for selecting an individual player's data
    def playerSelect(self, selectedPlayer):
               
        if(selectedPlayer == "brandonChan"):
            self.buttonReset()
            self.brandonChan = Button(self.playerSelection, text="Chan", command=lambda:self.buttonChange("brandonChan"), bg="blue", height=4, width=15)
            self.brandonChan.grid(row=0, column=0)
        
        if(selectedPlayer == "callumAshton"):
            self.buttonReset()
            self.callumAshton = Button(self.playerSelection, text="Callum", command=lambda:self.buttonChange("callumAshton"), bg="blue", height=4, width=15)
            self.callumAshton.grid(row=0, column=1)
            
        
        if(selectedPlayer == "danielPark"):
            self.buttonReset()
            self.danielPark = Button(self.playerSelection, text="Daniel", command=lambda:self.buttonChange("danielPark"), bg="blue", height=4, width=15)   
            self.danielPark.grid(row=0, column=2)
        
        
        if(selectedPlayer == "deirdreTruong"):
            self.buttonReset()
            self.deirdreTruong = Button(self.playerSelection, text="Deirdre", command=lambda:self.buttonChange("deirdreTruong"), bg="blue", height=4, width=15)
            self.deirdreTruong.grid(row=0, column=3)
        
        if(selectedPlayer == "edwardKang"):
            self.buttonReset()
            self.edwardKang = Button(self.playerSelection, text="Edward", command=lambda:self.buttonChange("edwardKang"), bg="blue", height=4, width=15)
            self.edwardKang.grid(row=0, column=4)
        
        if(selectedPlayer == "kevinMa"):
            self.buttonReset()
            self.kevinMa = Button(self.playerSelection, text="Kema", command=lambda:self.buttonChange("kevinMa"), bg="blue", height=4, width=15)
            self.kevinMa.grid(row=1, column=0)
        
        
        if(selectedPlayer == "kevinTang"):
            self.buttonReset()
            self.kevinTang = Button(self.playerSelection, text="Ktang", command=lambda:self.buttonChange("kevinTang"), bg="blue", height=4, width=15)
            self.kevinTang.grid(row=1, column=1)
        
        if(selectedPlayer == "lachlanDenham"):
            self.buttonReset()
            self.lachlanDenham = Button(self.playerSelection, text="Lachlan", command=lambda:self.buttonChange("lachlanDenham"), bg="blue", height=4, width=15)
            self.lachlanDenham.grid(row=1, column=2)
    
        
        if(selectedPlayer == "mimiChen"):
            self.buttonReset()
            self.mimiChen = Button(self.playerSelection, text="Mimi", command=lambda:self.buttonChange("mimiChen"), bg="blue", height=4, width=15)
            self.mimiChen.grid(row=1, column=3)
            
        
        if(selectedPlayer == "willOuyang"):
            self.buttonReset()
            self.willOuyang = Button(self.playerSelection, text="Will", command=lambda:self.buttonChange("willOuyang"), bg="blue", height=4, width=15)
            self.willOuyang.grid(row=1, column=4)
        
    #Function for reseting existing player selection upon new selection
    def buttonReset(self):
        
        self.brandonChan = Button(self.playerSelection, text="Chan", command=lambda: self.playerSelect("brandonChan"), height=4, width=15)
        self.callumAshton = Button(self.playerSelection, text="Callum", command=lambda: self.playerSelect("callumAshton"), height=4, width=15)
        self.danielPark = Button(self.playerSelection, text="Daniel", command=lambda: self.playerSelect("danielPark"), height=4, width=15)
        self.deirdreTruong = Button(self.playerSelection, text="Deirdre", command=lambda: self.playerSelect("deirdreTruong"), height=4, width=15)
        self.edwardKang = Button(self.playerSelection, text="Edward", command=lambda: self.playerSelect("edwardKang"), height=4, width=15)
        self.kevinMa = Button(self.playerSelection, text="Kema", command=lambda: self.playerSelect("kevinMa"), height=4, width=15)
        self.kevinTang = Button(self.playerSelection, text="Ktang", command=lambda: self.playerSelect("kevinTang"), height=4, width=15)
        self.lachlanDenham = Button(self.playerSelection, text="Lachlan", command=lambda: self.playerSelect("lachlanDenham"), height=4, width=15)
        self.mimiChen = Button(self.playerSelection, text="Mimi", command=lambda: self.playerSelect("mimiChen"), height=4, width=15)
        self.willOuyang = Button(self.playerSelection, text="Will", command=lambda: self.playerSelect("willOuyang"), height=4, width=15)
        
        self.brandonChan.grid(row=0, column=0)
        self.callumAshton.grid(row=0, column=1)
        self.danielPark.grid(row=0, column=2)
        self.deirdreTruong.grid(row=0, column=3)
        self.edwardKang.grid(row=0, column=4)
        self.kevinMa.grid(row=1, column=0)
        self.kevinTang.grid(row=1, column=1)
        self.lachlanDenham.grid(row=1, column=2)
        self.mimiChen.grid(row=1, column=3)
        self.willOuyang.grid(row=1, column=4)
    
    
    def buttonChange(self, playerName):
        self.selectedPlayer = playerName
        return
    
    def statIncrease(self, getStatType):
        playerList = ["brandonChan", "callumAshton", "danielPark", "deirdreTruong", "edwardKang", "kevinMa", "kevinTang", "lachlanDenham", "mimiChen", "willOuyang"]
        statTypeList = ["Serve Error", "Serve Success", "Receive Errors", "Receive Passes", "Spike Errors", "Spike Success", "Block Errors", "Block Successes", "Faults"]
        columnList = ['B', 'C', 'E', 'F', 'H', 'I', 'K', 'L', 'N']
        columnChar = columnList[statTypeList.index(getStatType)]
        playerRow = playerList.index(self.selectedPlayer) + 1
        weekRowSelect = (self.weekNumber-1 * 13)
        rowNumber = playerRow + weekRowSelect + 2
        self.ws[(columnChar + str(playerRow))] = self.ws[(columnChar + str(playerRow))].value + 1
        return

    def statDecrease(self, getStatType):
        playerList = ["brandonChan", "callumAshton", "danielPark", "deirdreTruong", "edwardKang", "kevinMa", "kevinTang", "lachlanDenham", "mimiChen", "willOuyang"]
        statTypeList = ["Serve Error", "Serve Success", "Receive Errors", "Receive Passes", "Spike Errors", "Spike Success", "Block Errors", "Block Successes", "Faults"]
        columnList = ['B', 'C', 'E', 'F', 'H', 'I', 'K', 'L', 'N']
        columnChar = columnList[statTypeList.index(getStatType)]
        playerRow = playerList.index(self.selectedPlayer) + 1
        weekRowSelect = (self.weekNumber-1 * 13)
        rowNumber = playerRow + weekRowSelect + 2
        if(self.ws[(columnChar + str(playerRow))].value >= 1):
            self.ws[(columnChar + str(playerRow))] = self.ws[(columnChar + str(playerRow))].value - 1
        else:
            messagebox.showinfo("Error", "You Cannot Decrease This Value Below 0")
        return