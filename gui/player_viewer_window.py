from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from .stats_editor import statsEditor
from tkinter import messagebox
from openpyxl.styles import Alignment


class playerViewerWindow():

    def __init__(self, master):
        master.title("Player Statistics")
        
        self.titleLabel = Label(master, text="Player Statistics")

        self.wb = Workbook()
        self.wb = load_workbook('data/volley_stats.xlsx')

        self.ws = self.wb['Team Info']

        self.titleLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))

        return
