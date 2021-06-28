from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from .stats_editor import statsEditor
from tkinter import messagebox

class seasonSelectionWindow():

    def __init__(self, master):
        master.title('Season Selector')
        #master.geometry("300x300")
        
        self.titleLabel = Label(master, text="Season Selector")

        self.wb = Workbook()
        self.wb = load_workbook('data/volley_stats.xlsx')

        self.ws = self.wb['Team Info']


        self.seasonCount = self.ws[('A2')].value
        self.seasonList = []

        for i in range(self.seasonCount):
            self.seasonList.append("Season " + str(i+1))
        
        self.dropDownSelection = StringVar()
        self.dropDownSelection.set(str(self.seasonList[0]))

        self.dropDownMenu = OptionMenu(master, self.dropDownSelection, *self.seasonList)

        self.acceptButton = Button(master, text= "Select", command=lambda: self.selectSeason(), padx=20, pady=10)
        self.exitButton = Button(master, text= "Cancel", command=master.destroy, padx=20, pady=10)
        
        self.newSeasonFrame = LabelFrame(master, text="Add/Remove Season", padx=10, pady=10)
        self.seasonAdd = Button(self.newSeasonFrame, text="+", command=lambda: self.addSeasonWindow(master), height=4, width=10)
        self.seasonRemove = Button(self.newSeasonFrame, text="-", command=lambda: self.removeSeasonWindow(master), height=4, width=10)

        self.titleLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        self.dropDownMenu.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        self.newSeasonFrame.grid(row=2, column=0, columnspan=2, padx=20)
        self.seasonAdd.grid(row=0, column=0)
        self.seasonRemove.grid(row=0, column=1)

        self.acceptButton.grid(row=3, column=0, padx=10, pady=10)
        self.exitButton.grid(row=3, column=1, padx=10, pady=10)


        return

    def addSeasonWindow(self, master):
        top = Toplevel()
        top.title("Add New Season")
        promptLabel = Label(top, text='How many weeks is this new Season?')
        inputField = Entry(top, width=20)
        sureButton = Button(top, text= "OK", command=lambda:self.addSeason(master, inputField, top), padx=20, pady=10)
        cancelButton = Button(top, text= "Cancel", command=lambda:top.destroy(), padx=20, pady=10)
        
        promptLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        inputField.grid(row=1, column=0, columnspan=2)
        sureButton.grid(row=2, column=0, padx=10, pady=10)
        cancelButton.grid(row=2, column=1, padx=10, pady=10)
        return

    def addSeason(self, master, inputField, top):
        weekNo = inputField.get()
        cellRangeList = []
        if(not(weekNo.isdigit())):
            messagebox.showinfo("Error", "Enter a valid number of Weeks")
            return
        self.ws[('A' + str(((len(self.seasonList))*5) + 3))].value = 'Season No.'
        self.ws[('A' + str(((len(self.seasonList))*5) + 4))].value = len(self.seasonList) + 1
        self.ws[('B' + str(((len(self.seasonList))*5) + 3))].value = 'Player Count'
        self.ws[('B' + str(((len(self.seasonList))*5) + 4))].value = 0
        self.ws[('C' + str(((len(self.seasonList))*5) + 3))].value = 'Number of Weeks'
        self.ws[('C' + str(((len(self.seasonList))*5) + 4))].value = int(weekNo)
        self.ws[('A' + str(((len(self.seasonList))*5) + 5))].value = 'Player Names'
        ws2 = self.wb.create_sheet("Season " + str(len(self.seasonList) + 1))
        cellRange = self.ws['K2':'AD2']
        for cell in cellRange:
            for x in cell:
                cellRangeList.append(x.value)
        for j in range(int(weekNo)):
            ws2['A' + str((int(j)*3)+1)].value = 'Week ' + str(j+1)
            for i in ws2['A' + str((j*3)+2):'T' + str((j*3)+2)]:
                for k in i:
                    k.value = cellRangeList[i.index(k)]
        self.ws[('A2')].value = self.ws[('A2')].value + 1
        self.seasonCount = self.ws[('A2')].value
        self.seasonList.clear()
        for i in range(self.seasonCount):
            self.seasonList.append("Season " + str(i+1))
        self.dropDownMenu = OptionMenu(master, self.dropDownSelection, *self.seasonList)
        self.dropDownMenu.grid_forget()
        self.dropDownMenu.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        self.dropDownSelection.set(str(self.seasonList[0]))
        self.wb.save('data/volley_stats.xlsx')
        messagebox.showinfo("Success", "Season " + str(len(self.seasonList)) + " was added.")
        top.destroy()
        return

    def removeSeasonWindow(self, master):
        top = Toplevel()
        top.title("Remove Season")
        warningLabel = Label(top, text='Are you sure you want to delete Season ' + str(self.dropDownSelection.get() + "?"))
        sureButton = Button(top, text= "OK", command=lambda:self.deleteSeason(top, master), padx=20, pady=10)
        cancelButton = Button(top, text= "Cancel", command=lambda:top.destroy(), padx=20, pady=10)
        
        warningLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        sureButton.grid(row=1, column=0, padx=10, pady=10)
        cancelButton.grid(row=1, column=1, padx=10, pady=10)

        return

    def deleteSeason(self, top, master):
        self.wb.remove_sheet(self.wb.get_sheet_by_name(self.dropDownSelection.get()))
        self.ws[('A2')].value = self.ws[('A2')].value - 1
        self.seasonCount = self.ws[('A2')].value
        self.ws.delete_rows((self.seasonList.index(self.dropDownSelection.get())*5) + 3, 5)
        for i in range(self.seasonList.index(self.dropDownSelection.get()) + 1 , len(self.seasonList)):
            self.ws[('A' + str(((i - 1) * 5) + 4))].value = self.ws[('A' + str(((i - 1) * 5) + 4))].value - 1

        self.wb.save('data/volley_stats.xlsx')
        if((self.seasonList.index(self.dropDownSelection.get()) + 1) < len(self.seasonList)):
            for i in range((self.seasonList.index(self.dropDownSelection.get()) + 1), len(self.seasonList)):
                ws1 = self.wb['Season ' + str(i+1)]
                ws1.title = 'Season ' + str(i)
            self.wb.save('data/volley_stats.xlsx')
        self.seasonList.clear()
        for i in range(self.seasonCount):
            self.seasonList.append("Season " + str(i+1))
        self.dropDownMenu = OptionMenu(master, self.dropDownSelection, *self.seasonList)
        self.dropDownMenu.grid_forget()
        self.dropDownMenu.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        messagebox.showinfo("Success", str(self.dropDownSelection.get()) + " was deleted.")
        self.dropDownSelection.set(str(self.seasonList[0]))
        top.destroy()
        return


    def selectSeason(self):
        seasonSelection = self.dropDownSelection.get()
        top = Toplevel()
        b = statsEditor(top, seasonSelection, (self.seasonList.index(self.dropDownSelection.get()) + 1))
        top.mainloop()
        return

