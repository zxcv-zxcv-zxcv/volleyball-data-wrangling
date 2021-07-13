from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
from .add_player_window import addPlayerWindow
from openpyxl.styles import Alignment

class statsEditor():
     
    
    def __init__(self, master, seasonName, seasonNo):
        
        
        self.wb = Workbook()
        self.wb = load_workbook('data/volley_stats.xlsx')

        self.ws1 = self.wb[seasonName]
        
        self.ws2 = self.wb['Team Info']
        
        
        self.seasonNumber = seasonNo

        self.playerNumber = self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value
        
        self.playerList = []
        cellRange = self.ws2[str(((self.seasonNumber-1) * 7) + 6)]
        for i in range(self.playerNumber):
            self.playerList.append(str(cellRange[i].value))

        self.playerNicknameList = []
        cellRange = self.ws2[str(((self.seasonNumber-1) * 7) + 8)]
        for i in range(self.playerNumber):
            self.playerNicknameList.append("".join(str(cellRange[i].value).split()))

        self.weekList = []
        for i in range(self.ws2[('C' + str(((self.seasonNumber-1) * 7) + 4))].value):
            self.weekList.append(i+1)
        self.weekNumber = 1

        self.buttonList = []
        for i in range(len(self.playerList)):
            self.buttonList.append(str(self.playerList[i]))
        
        self.selectedPlayer = "None"
        self.toggleBool = False


        master.title('Volleyball Statistics Input')

        teamName = self.ws2[('C2')].value
        self.titleLabel = Label(master, text=teamName + "\n" + seasonName)

        #Initializing add/remove player Buttons
        self.newPlayerFrame = LabelFrame(master, text="Change Players", padx=5, pady=5)
        self.playerAdd = Button(self.newPlayerFrame, text="+", command=lambda: self.addPlayerWindow(), height=2, width=5)
        self.playerRemove = Button(self.newPlayerFrame, text="-", command=lambda: self.removePlayerWindow(), height=2, width=5)

        #Initializing Week Selection GUI
        self.weekFrame = LabelFrame(master, text="Week Selection", padx=10, pady=10)
        self.weekLabel = Label(self.weekFrame, text="Week: 1 of "+ str(len(self.weekList)), padx=20, pady=10)
        self.prevWeekButton = Button(self.weekFrame, text= "<<", command=lambda: self.prevWeek(), padx=10, pady=10, anchor=W) #
        self.nextWeekButton = Button(self.weekFrame, text= ">>", command=lambda: self.nextWeek(), padx=10, pady=10, anchor=W) 
    
        #Initializing Player Selection Buttons
        self.playerSelection = LabelFrame(master, text="Player Selection", padx=5, pady=5)
        for i in range(len(self.playerList)):
            self.buttonList[i] = Button(self.playerSelection, text=self.playerNicknameList[i], command=lambda x=i: self.playerSelect(x), height=4, width=15)
            self.buttonList[i].grid(row=int(i//5), column=(i%5))

        #Initializing Statisitic Add and Subtract Buttons
        self.servesFrame = LabelFrame(master, text="Serving", padx=5, pady=5)
        self.serveAceButton = Button(self.servesFrame, text="Ace", command=lambda: self.statIncrease("serveAce"), height=3, width=10)
        self.serveInButton = Button(self.servesFrame, text="In", command=lambda: self.statIncrease("serveIn"), height=3, width=10)
        self.serveOutButton = Button(self.servesFrame, text="Out", command=lambda: self.statIncrease("serveOut"), height=3, width=10)
        self.serveShortButton = Button(self.servesFrame, text="Short", command=lambda: self.statIncrease("serveShort"), height=3, width=10)
        

        self.receivesFrame = LabelFrame(master, text="Receiving", padx=5, pady=5)
        self.receiveTargetedButton = Button(self.receivesFrame, text="Targeted", command=lambda: self.statIncrease("receiveTargeted"), height=3, width=10)
        self.receiveHighButton = Button(self.receivesFrame, text="High", command=lambda: self.statIncrease("receiveHigh"), height=3, width=10)
        self.receiveOffButton = Button(self.receivesFrame, text="Off", command=lambda: self.statIncrease("receiveOff"), height=3, width=10)
        self.receiveLowButton = Button(self.receivesFrame, text="Low", command=lambda: self.statIncrease("receiveLow"), height=3, width=10)


        self.setsFrame = LabelFrame(master, text="Setting", padx=5, pady=5)
        self.setTargetedButton = Button(self.setsFrame, text="Targeted", command=lambda: self.statIncrease("setTargeted"), height=3, width=10)
        self.setHighButton = Button(self.setsFrame, text="High", command=lambda: self.statIncrease("setHigh"), height=3, width=10)
        self.setOffButton = Button(self.setsFrame, text="Off", command=lambda: self.statIncrease("setOff"), height=3, width=10)
        self.setLowButton = Button(self.setsFrame, text="Low", command=lambda: self.statIncrease("setLow"), height=3, width=10)


        self.spikesFrame = LabelFrame(master, text="Spiking", padx=5, pady=5)
        self.spikeSuccessButton = Button(self.spikesFrame, text="Success", command=lambda: self.statIncrease("spikeSuccess"), height=3, width=10)
        self.spikeInButton = Button(self.spikesFrame, text="In", command=lambda: self.statIncrease("spikeIn"), height=3, width=10)
        self.spikeBlockedButton = Button(self.spikesFrame, text="Blocked", command=lambda: self.statIncrease("spikeBlocked"), height=3, width=10)
        self.spikeOutButton = Button(self.spikesFrame, text="Out", command=lambda: self.statIncrease("spikeOut"), height=3, width=10)
        

        self.tipsFrame = LabelFrame(master, text="Tipping", padx=5, pady=5)
        self.tipSuccessButton = Button(self.tipsFrame, text="Success", command=lambda: self.statIncrease("tipSuccess"), height=3, width=10)
        self.tipInButton = Button(self.tipsFrame, text="In", command=lambda: self.statIncrease("tipIn"), height=3, width=10)
        self.tipBlockedButton = Button(self.tipsFrame, text="Blocked", command=lambda: self.statIncrease("tipBlocked"), height=3, width=10)
        self.tipOutButton = Button(self.tipsFrame, text="Out", command=lambda: self.statIncrease("tipOut"), height=3, width=10)


        self.blocksFrame = LabelFrame(master, text="Blocking", padx=5, pady=5)
        self.blockScoreButton = Button(self.blocksFrame, text="Score", command=lambda: self.statIncrease("blockScore"), height=3, width=10)
        self.blockTouchButton = Button(self.blocksFrame, text="Touch", command=lambda: self.statIncrease("blockTouch"), height=3, width=10)
        self.blockOffButton = Button(self.blocksFrame, text="Off", command=lambda: self.statIncrease("blockOff"), height=3, width=10)
        self.blockFailButton = Button(self.blocksFrame, text="No Block", command=lambda: self.statIncrease("blockFail"), height=3, width=10)


        self.Faults = LabelFrame(master, text="Faults", padx=5, pady=5)
        self.FaultsAdd = Button(self.Faults, text="+", command=lambda: self.statIncrease("Faults"), height=2, width=5)
        self.FaultsRemove = Button(self.Faults, text="-", command=lambda: self.statDecrease("Faults"), height=2, width=5)
        self.FaultsLabel = Label(self.Faults, text="Faults: ", padx=10)

        self.toggleButton = Button(master, text="Toggle \n Decrease", command=lambda:self.toggleStatButtons(master), height=3, width=10)
        
        #Initialising Statistics table
        self.statisticsmaster = LabelFrame(master, text="Player Statistics", padx=10, pady=7)

        
        self.serveLabelFrame = Frame(master)
        self.serveAceLabel = Label(self.serveLabelFrame, text="Aces: ", padx=10)
        self.serveInLabel = Label(self.serveLabelFrame, text="In: ", padx=10)
        self.serveOutLabel = Label(self.serveLabelFrame, text="Out: ", padx=10)
        self.serveShortLabel = Label(self.serveLabelFrame, text="Short: ", padx=10)
        self.serveRateLabel = Label(self.serveLabelFrame, text="Serve Rate: ", padx=10)
        self.serveAceRateLabel  = Label(self.serveLabelFrame, text="Ace Rate: ", padx=10)

        self.receiveLabelFrame = Frame(master)
        self.receiveTargetedLabel = Label(self.receiveLabelFrame, text="Targeted: ", padx=10)
        self.receiveHighLabel = Label(self.receiveLabelFrame, text="High: ", padx=10)
        self.receiveOffLabel = Label(self.receiveLabelFrame, text="Off: ", padx=10)
        self.receiveLowLabel = Label(self.receiveLabelFrame, text="Low: ", padx=10)
        self.receiveRateLabel = Label(self.receiveLabelFrame, text="Pass Rate: ", padx=10)
        self.receiveBestLabel = Label(self.receiveLabelFrame, text="Perfect Rate: ", padx=10)

        self.setLabelFrame = Frame(master)
        self.setTargetedLabel = Label(self.setLabelFrame, text="Targeted: ", padx=10)
        self.setHighLabel = Label(self.setLabelFrame, text="High: ", padx=10)
        self.setOffLabel = Label(self.setLabelFrame, text="Off: ", padx=10)
        self.setLowLabel = Label(self.setLabelFrame, text="Low: ", padx=10)
        self.setRateLabel = Label(self.setLabelFrame, text="Set Rate: ", padx=10)
        self.setBestLabel = Label(self.setLabelFrame, text="Perfect Rate: ", padx=10)

        self.spikeLabelFrame = Frame(master)
        self.spikeSuccessLabel = Label(self.spikeLabelFrame, text="Score: ", padx=10)
        self.spikeInLabel = Label(self.spikeLabelFrame, text="In: ", padx=10)
        self.spikeBlockedLabel = Label(self.spikeLabelFrame, text="Blocked: ", padx=10)
        self.spikeOutLabel = Label(self.spikeLabelFrame, text="Out: ", padx=10)
        self.spikeRateLabel = Label(self.spikeLabelFrame, text="Spike Rate: ", padx=10)
        self.spikeScoreRateLabel = Label(self.spikeLabelFrame, text="Score Rate: ", padx=10)

        self.tipLabelFrame = Frame(master)
        self.tipSuccessLabel = Label(self.tipLabelFrame, text="Score: ", padx=10)
        self.tipInLabel = Label(self.tipLabelFrame, text="In: ", padx=10)
        self.tipBlockedLabel = Label(self.tipLabelFrame, text="Blocked: ", padx=10)
        self.tipOutLabel = Label(self.tipLabelFrame, text="Out: ", padx=10)
        self.tipRateLabel = Label(self.tipLabelFrame, text="Tip Rate: ", padx=10)
        self.tipScoreLabel = Label(self.tipLabelFrame, text="Score Rate: ", padx=10)

        self.blockLabelFrame = Frame(master)
        self.blockScoreLabel = Label(self.blockLabelFrame, text="Score: ", padx=10)
        self.blockTouchLabel = Label(self.blockLabelFrame, text="Touch: ", padx=10)
        self.blockOffLabel = Label(self.blockLabelFrame, text="Off: ", padx=10)
        self.blockFailLabel = Label(self.blockLabelFrame, text="No block: ", padx=10)
        self.blockRateLabel = Label(self.blockLabelFrame, text="Block Rate: ", padx=10)
        self.blockScoreRateLabel = Label(self.blockLabelFrame, text="Score Rate: ", padx=10)

        ##Attaching all initial state GUI components to grid
        
        #Initial elements
        self.titleLabel.grid(row=0, column=0, columnspan=2)

        self.newPlayerFrame.grid(row=0, column=4, columnspan=2)
        self.playerAdd.grid(row=0, column=0)
        self.playerRemove.grid(row=0, column=1)
        
        self.weekFrame.grid(row=0, column=6, columnspan=2, padx=(0,20))
        self.prevWeekButton.grid(row=0, column=0)
        self.weekLabel.grid(row=0, column=1)
        self.nextWeekButton.grid(row=0, column=2)
     
        #Player Selection buttons
        self.playerSelection.grid(row=1, column=2, columnspan=6, rowspan=2, padx=10, pady=5)
      
        #buttons and frames for changing data
        self.servesFrame.grid(row=3, column=0, columnspan=4, padx=10, pady=5)
        self.serveAceButton.grid(row=0, column=0)
        self.serveInButton.grid(row=0, column=1)
        self.serveOutButton.grid(row=0, column=2)
        self.serveShortButton.grid(row=0, column=3)


        self.receivesFrame.grid(row=3, column=4, columnspan=4, padx=10, pady=5)
        self.receiveTargetedButton.grid(row=0, column=0)
        self.receiveHighButton.grid(row=0, column=1)
        self.receiveOffButton.grid(row=0, column=2)
        self.receiveLowButton.grid(row=0, column=3)


        self.setsFrame.grid(row=5, column=0, columnspan=4, padx=10, pady=5)
        self.setTargetedButton.grid(row=0, column=0)
        self.setHighButton.grid(row=0, column=1)
        self.setOffButton.grid(row=0, column=2)
        self.setLowButton.grid(row=0, column=3)


        self.spikesFrame.grid(row=5, column=4, columnspan=4, padx=10, pady=5)
        self.spikeSuccessButton.grid(row=0, column=0)
        self.spikeInButton.grid(row=0, column=1)
        self.spikeBlockedButton.grid(row=0, column=2)
        self.spikeOutButton.grid(row=0, column=3)


        self.tipsFrame.grid(row=7, column=0, columnspan=4, padx=10, pady=5)
        self.tipSuccessButton.grid(row=0, column=0)
        self.tipInButton.grid(row=0, column=1)
        self.tipBlockedButton.grid(row=0, column=2)
        self.tipOutButton.grid(row=0, column=3)


        self.blocksFrame.grid(row=7, column=4, columnspan=4, padx=10, pady=5)
        self.blockScoreButton.grid(row=0, column=0)
        self.blockTouchButton.grid(row=0, column=1)
        self.blockOffButton.grid(row=0, column=2)
        self.blockFailButton.grid(row=0, column=3)


        self.Faults.grid(row=1, column=0, columnspan=2, padx=10)
        self.FaultsAdd.grid(row=0, column=0)
        self.FaultsRemove.grid(row=0, column=1)
        self.FaultsLabel.grid(row=1, column=0, columnspan=2, pady=4)
        
        self.toggleButton.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

        #Statistics frame and labels
        self.serveLabelFrame.grid(row=4, column=0, columnspan=4)
        self.serveAceLabel.grid(row=4, column=0, pady=5)
        self.serveInLabel.grid(row=4, column=1, pady=5)
        self.serveOutLabel.grid(row=4, column=2, pady=5)
        self.serveShortLabel.grid(row=4, column=3, pady=5)
        
        self.receiveLabelFrame.grid(row=4, column=4, columnspan=4)
        self.receiveTargetedLabel.grid(row=4, column=0, pady=5)
        self.receiveHighLabel.grid(row=4, column=1, pady=5)
        self.receiveOffLabel.grid(row=4, column=2, pady=5)
        self.receiveLowLabel.grid(row=4, column=3, pady=5)
        
        self.setLabelFrame.grid(row=6, column=0, columnspan=4)
        self.setTargetedLabel.grid(row=6, column=0, pady=5)
        self.setHighLabel.grid(row=6, column=1, pady=5)
        self.setOffLabel.grid(row=6, column=2, pady=5)
        self.setLowLabel.grid(row=6, column=3, pady=5)
        
        self.spikeLabelFrame.grid(row=6, column=4, columnspan=4)
        self.spikeSuccessLabel.grid(row=6, column=0, pady=5)
        self.spikeInLabel.grid(row=6, column=1, pady=5)
        self.spikeBlockedLabel.grid(row=6, column=2, pady=5)
        self.spikeOutLabel.grid(row=6, column=3, pady=5)
        
        self.tipLabelFrame.grid(row=8, column=0, columnspan=4)
        self.tipSuccessLabel.grid(row=8, column=0, pady=(5, 15))
        self.tipInLabel.grid(row=8, column=1, pady=(5, 15))
        self.tipBlockedLabel.grid(row=8, column=2, pady=(5, 15))
        self.tipOutLabel.grid(row=8, column=3, pady=(5, 15))
        
        self.blockLabelFrame.grid(row=8, column=4, columnspan=4)
        self.blockScoreLabel.grid(row=8, column=0, pady=(5, 15))
        self.blockTouchLabel.grid(row=8, column=1, pady=(5, 15))
        self.blockOffLabel.grid(row=8, column=2, pady=(5, 15))
        self.blockFailLabel.grid(row=8, column=3, pady=(5, 15))
        

    #Function for iterating week forwards once
    def nextWeek(self): 
        self.weekLabel.grid_forget()
        
        if(self.weekNumber >= len(self.weekList)):
            self.weekNumber = 1
        else:
            self.weekNumber += 1
        self.weekLabel = Label(self.weekFrame, text="Week: " + str(self.weekList[self.weekNumber-1]) + " of " + str(len(self.weekList)), padx=20, pady=10, anchor=W)
        self.weekLabel.grid(row=0, column=1)
        if(self.selectedPlayer != "None"):
            self.updateStatsLabels()
        return
    
    #Function for iterating week backwards once
    def prevWeek(self):
      
        self.weekLabel.grid_forget()
        
        if(self.weekNumber-1 == 0):
            self.weekNumber = len(self.weekList)
        else:
            self.weekNumber -= 1 
        self.weekLabel = Label(self.weekFrame, text="Week: " + str(self.weekList[self.weekNumber-1]) + " of " + str(len(self.weekList)), padx=20, pady=10, anchor=W)
        self.weekLabel.grid(row=0, column=1)
        if(self.selectedPlayer != "None"):
            self.updateStatsLabels()
        return
    
    #Function for selecting an individual player's data
    def playerSelect(self, playerName):
        self.buttonList[playerName] = Button(self.playerSelection, text=self.playerNicknameList[playerName], command=lambda x=playerName: self.playerSelect(x), bg="blue", height=4, width=15)
        self.buttonList[playerName].grid(row=int(playerName//5), column=(playerName%5))
        if(self.selectedPlayer != "None"):
            self.buttonList[self.playerList.index(self.selectedPlayer)] = Button(self.playerSelection, text=self.playerNicknameList[self.playerList.index(self.selectedPlayer)], command=lambda x=self.playerList.index(self.selectedPlayer): self.playerSelect(x), height=4, width=15)
            self.buttonList[self.playerList.index(self.selectedPlayer)].grid(row=int(self.playerList.index(self.selectedPlayer)//5), column=self.playerList.index(self.selectedPlayer)%5)
        self.selectedPlayer = self.playerList[playerName]
        self.updateStatsLabels()   
        return
        
    #Function for reseting existing player selection upon new selection
    def buttonReset(self):
        for i in range(len(self.buttonList)):
            self.buttonList[i] = Button(self.playerSelection, text=self.playerNicknameList[i], command=lambda x=i: self.playerSelect(x), height=4, width=15)
            self.buttonList[i].grid(row=int(i//5), column=(i%5))
        self.selectedPlayer = "None"
        
    
    def statIncrease(self, getStatType):
        statTypeList = ["serveAce", "serveIn", "serveOut", "serveShort", "receiveTargeted", "receiveHigh", "receiveOff", "receiveLow", "setTargeted", "setHigh", "setOff", "setLow", "spikeSuccess", "spikeIn", "spikeBlocked", "spikeOut", "tipSuccess", "tipIn", "tipBlocked", "tipOut", "blockScore", "blockTouch", "blockOff", "blockFail", "Faults"]
        columnList = ['B', 'C', 'D', 'E', 'H', 'I', 'J', 'K', 'N', 'O', 'P', 'Q', 'T', 'U', 'V', 'W', 'Z', 'AA', 'AB', 'AC', 'AF', 'AG', 'AH', 'AI', 'AL']
        columnChar = columnList[statTypeList.index(getStatType)]
        if(self.selectedPlayer == "None"):
            messagebox.showinfo("Error", "No Player Selected")
            return
        
        rowNumber = self.getRowNumber()
        self.ws1[(columnChar + str(rowNumber))] = self.ws1[(columnChar + str(rowNumber))].value + 1
        self.updateStatsLabels()
        return

    def statDecrease(self, getStatType):
        statTypeList = ["serveAce", "serveIn", "serveOut", "serveShort", "receiveTargeted", "receiveHigh", "receiveOff", "receiveLow", "setTargeted", "setHigh", "setOff", "setLow", "spikeSuccess", "spikeIn", "spikeBlocked", "spikeOut", "tipSuccess", "tipIn", "tipBlocked", "tipOut", "blockScore", "blockTouch", "blockOff", "blockFail", "Faults"]
        columnList = ['B', 'C', 'D', 'E', 'H', 'I', 'J', 'K', 'N', 'O', 'P', 'Q', 'T', 'U', 'V', 'W', 'Z', 'AA', 'AB', 'AC', 'AF', 'AG', 'AH', 'AI', 'AL']
        columnChar = columnList[statTypeList.index(getStatType)]
        if(self.selectedPlayer == "None"):
            messagebox.showinfo("Error", "No Player Selected")
            return
        
        rowNumber = self.getRowNumber()
        if(self.ws1[(columnChar + str(rowNumber))].value >= 1):
            self.ws1[(columnChar + str(rowNumber))] = self.ws1[(columnChar + str(rowNumber))].value - 1
            self.updateStatsLabels()
        else:
            messagebox.showinfo("Error", "You Cannot Decrease This Value Below 0")
        self.updateStatsLabels()
        return

    def getRowNumber(self):
        playerRow = self.playerList.index(self.selectedPlayer) + 1
        weekRowSelect = ((self.weekNumber-1) * (len(self.playerList) + 3))
        rowNumber = playerRow + weekRowSelect + 2
        return rowNumber


    def updateStatsLabels(self):
        
        rowNumber = self.getRowNumber()

        self.serveAceLabel.grid_forget()
        self.serveInLabel.grid_forget()
        self.serveOutLabel.grid_forget()
        self.serveShortLabel.grid_forget()
        
        self.receiveTargetedLabel.grid_forget()
        self.receiveHighLabel.grid_forget()
        self.receiveOffLabel.grid_forget()
        self.receiveLowLabel.grid_forget()
        
        self.setTargetedLabel.grid_forget()
        self.setHighLabel.grid_forget()
        self.setOffLabel.grid_forget()
        self.setLowLabel.grid_forget()
        
        self.spikeSuccessLabel.grid_forget()
        self.spikeInLabel.grid_forget()
        self.spikeBlockedLabel.grid_forget()
        self.spikeOutLabel.grid_forget()
        
        self.tipSuccessLabel.grid_forget()
        self.tipInLabel.grid_forget()
        self.tipBlockedLabel.grid_forget()
        self.tipOutLabel.grid_forget()
        
        self.blockScoreLabel.grid_forget()
        self.blockTouchLabel.grid_forget()
        self.blockOffLabel.grid_forget()
        self.blockFailLabel.grid_forget()
        
        if((self.ws1['B' + str(rowNumber)].value + self.ws1['C' + str(rowNumber)].value + self.ws1['D' + str(rowNumber)].value + self.ws1['E' + str(rowNumber)].value) != 0):
            self.ws1['F' + str(rowNumber)].value = str(round((self.ws1['B' + str(rowNumber)].value + self.ws1['C' + str(rowNumber)].value) / (self.ws1['B' + str(rowNumber)].value + self.ws1['C' + str(rowNumber)].value + self.ws1['D' + str(rowNumber)].value + self.ws1['E' + str(rowNumber)].value)*100)) + "%"
            self.ws1['G' + str(rowNumber)].value = str(round((self.ws1['B' + str(rowNumber)].value) / (self.ws1['B' + str(rowNumber)].value + self.ws1['C' + str(rowNumber)].value + self.ws1['D' + str(rowNumber)].value + self.ws1['E' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['F' + str(rowNumber)].value = "0%"
            self.ws1['G' + str(rowNumber)].value = "0%"
        
       
        if((self.ws1['H' + str(rowNumber)].value + self.ws1['I' + str(rowNumber)].value + self.ws1['J' + str(rowNumber)].value + self.ws1['K' + str(rowNumber)].value) != 0):
            self.ws1['L' + str(rowNumber)].value = str(round((self.ws1['H' + str(rowNumber)].value + self.ws1['I' + str(rowNumber)].value) / (self.ws1['H' + str(rowNumber)].value + self.ws1['I' + str(rowNumber)].value + self.ws1['J' + str(rowNumber)].value + self.ws1['K' + str(rowNumber)].value)*100)) + "%"
            self.ws1['M' + str(rowNumber)].value = str(round((self.ws1['H' + str(rowNumber)].value) / (self.ws1['H' + str(rowNumber)].value + self.ws1['I' + str(rowNumber)].value + self.ws1['J' + str(rowNumber)].value + self.ws1['K' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['L' + str(rowNumber)].value = "0%"
            self.ws1['M' + str(rowNumber)].value = "0%"
        

        if((self.ws1['N' + str(rowNumber)].value + self.ws1['O' + str(rowNumber)].value + self.ws1['P' + str(rowNumber)].value + self.ws1['Q' + str(rowNumber)].value) != 0):
            self.ws1['R' + str(rowNumber)].value = str(round((self.ws1['N' + str(rowNumber)].value + self.ws1['O' + str(rowNumber)].value) / (self.ws1['N' + str(rowNumber)].value + self.ws1['O' + str(rowNumber)].value + self.ws1['P' + str(rowNumber)].value + self.ws1['Q' + str(rowNumber)].value)*100)) + "%"
            self.ws1['S' + str(rowNumber)].value = str(round((self.ws1['N' + str(rowNumber)].value) / (self.ws1['N' + str(rowNumber)].value + self.ws1['O' + str(rowNumber)].value + self.ws1['P' + str(rowNumber)].value + self.ws1['Q' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['R' + str(rowNumber)].value = "0%"
            self.ws1['S' + str(rowNumber)].value = "0%"
        

        if((self.ws1['T' + str(rowNumber)].value + self.ws1['U' + str(rowNumber)].value + self.ws1['P' + str(rowNumber)].value + self.ws1['W' + str(rowNumber)].value) != 0):
            self.ws1['X' + str(rowNumber)].value = str(round((self.ws1['T' + str(rowNumber)].value + self.ws1['U' + str(rowNumber)].value) / (self.ws1['T' + str(rowNumber)].value + self.ws1['U' + str(rowNumber)].value + self.ws1['D' + str(rowNumber)].value + self.ws1['W' + str(rowNumber)].value)*100)) + "%"
            self.ws1['Y' + str(rowNumber)].value = str(round((self.ws1['T' + str(rowNumber)].value) / (self.ws1['T' + str(rowNumber)].value + self.ws1['U' + str(rowNumber)].value + self.ws1['D' + str(rowNumber)].value + self.ws1['W' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['X' + str(rowNumber)].value = "0%"
            self.ws1['Y' + str(rowNumber)].value = "0%"
        

        if((self.ws1['Z' + str(rowNumber)].value + self.ws1['AA' + str(rowNumber)].value + self.ws1['AB' + str(rowNumber)].value + self.ws1['AC' + str(rowNumber)].value) != 0):
            self.ws1['AD' + str(rowNumber)].value = str(round((self.ws1['Z' + str(rowNumber)].value + self.ws1['AA' + str(rowNumber)].value) / (self.ws1['Z' + str(rowNumber)].value + self.ws1['AA' + str(rowNumber)].value + self.ws1['AB' + str(rowNumber)].value + self.ws1['AC' + str(rowNumber)].value)*100)) + "%"
            self.ws1['AE' + str(rowNumber)].value = str(round((self.ws1['Z' + str(rowNumber)].value) / (self.ws1['Z' + str(rowNumber)].value + self.ws1['AA' + str(rowNumber)].value + self.ws1['AB' + str(rowNumber)].value + self.ws1['AC' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['AD' + str(rowNumber)].value = "0%"
            self.ws1['AE' + str(rowNumber)].value = "0%"
        

        if((self.ws1['AF' + str(rowNumber)].value + self.ws1['AG' + str(rowNumber)].value + self.ws1['AH' + str(rowNumber)].value + self.ws1['AI' + str(rowNumber)].value) != 0):
            self.ws1['AJ' + str(rowNumber)].value = str(round((self.ws1['AF' + str(rowNumber)].value + self.ws1['AG' + str(rowNumber)].value) / (self.ws1['AF' + str(rowNumber)].value + self.ws1['AG' + str(rowNumber)].value + self.ws1['AH' + str(rowNumber)].value + self.ws1['AI' + str(rowNumber)].value)*100)) + "%"
            self.ws1['AK' + str(rowNumber)].value = str(round((self.ws1['AF' + str(rowNumber)].value) / (self.ws1['AF' + str(rowNumber)].value + self.ws1['AG' + str(rowNumber)].value + self.ws1['AH' + str(rowNumber)].value + self.ws1['AI' + str(rowNumber)].value)*100)) + "%"    
        else:
            self.ws1['AJ' + str(rowNumber)].value = "0%"
            self.ws1['AK' + str(rowNumber)].value = "0%"
    

        self.serveAceLabel = Label(self.serveLabelFrame, text="Aces: " + str(self.ws1[('B' + str(rowNumber))].value), padx=10)
        self.serveInLabel = Label(self.serveLabelFrame, text="In: " + str(self.ws1[('C' + str(rowNumber))].value), padx=10)
        self.serveOutLabel = Label(self.serveLabelFrame, text="Out: " + str(self.ws1[('D' + str(rowNumber))].value), padx=10)
        self.serveShortLabel = Label(self.serveLabelFrame, text="Short: " + str(self.ws1[('E' + str(rowNumber))].value), padx=10)
        
        self.receiveTargetedLabel = Label(self.receiveLabelFrame, text="Targeted: " + str(self.ws1[('H' + str(rowNumber))].value), padx=10)
        self.receiveHighLabel = Label(self.receiveLabelFrame, text="High: " + str(self.ws1[('I' + str(rowNumber))].value), padx=10)
        self.receiveOffLabel = Label(self.receiveLabelFrame, text="Off: " + str(self.ws1[('J' + str(rowNumber))].value), padx=10)
        self.receiveLowLabel = Label(self.receiveLabelFrame, text="Low: " + str(self.ws1[('K' + str(rowNumber))].value), padx=10)
        
        self.setTargetedLabel = Label(self.setLabelFrame, text="Targeted: " + str(self.ws1[('N' + str(rowNumber))].value), padx=10)
        self.setHighLabel = Label(self.setLabelFrame, text="High: " + str(self.ws1[('O' + str(rowNumber))].value), padx=10)
        self.setOffLabel = Label(self.setLabelFrame, text="Off: " + str(self.ws1[('P' + str(rowNumber))].value), padx=10)
        self.setLowLabel = Label(self.setLabelFrame, text="Low: " + str(self.ws1[('Q' + str(rowNumber))].value), padx=10)
        
        self.spikeSuccessLabel = Label(self.spikeLabelFrame, text="Score: " + str(self.ws1[('T' + str(rowNumber))].value), padx=10)
        self.spikeInLabel = Label(self.spikeLabelFrame, text="In: " + str(self.ws1[('U' + str(rowNumber))].value), padx=10)
        self.spikeBlockedLabel = Label(self.spikeLabelFrame, text="Blocked: " + str(self.ws1[('V' + str(rowNumber))].value), padx=10)
        self.spikeOutLabel = Label(self.spikeLabelFrame, text="Out: " + str(self.ws1[('W' + str(rowNumber))].value), padx=10)
        
        self.tipSuccessLabel = Label(self.tipLabelFrame, text="Score: " + str(self.ws1[('Z' + str(rowNumber))].value), padx=10)
        self.tipInLabel = Label(self.tipLabelFrame, text="In: " + str(self.ws1[('AA' + str(rowNumber))].value), padx=10)
        self.tipBlockedLabel = Label(self.tipLabelFrame, text="Blocked: " + str(self.ws1[('AB' + str(rowNumber))].value), padx=10)
        self.tipOutLabel = Label(self.tipLabelFrame, text="Out: " + str(self.ws1[('AC' + str(rowNumber))].value), padx=10)
        
        self.blockScoreLabel = Label(self.blockLabelFrame, text="Score: " + str(self.ws1[('AF' + str(rowNumber))].value), padx=10)
        self.blockTouchLabel = Label(self.blockLabelFrame, text="Touch: " + str(self.ws1[('AG' + str(rowNumber))].value), padx=10)
        self.blockOffLabel = Label(self.blockLabelFrame, text="Off: " + str(self.ws1[('AH' + str(rowNumber))].value), padx=10)
        self.blockFailLabel = Label(self.blockLabelFrame, text="No block: " + str(self.ws1[('AI' + str(rowNumber))].value), padx=10)
        
        self.FaultsLabel = Label(self.Faults, text="Faults: " + str(self.ws1[('AL' + str(rowNumber))].value), padx=10)


        self.serveAceLabel.grid(row=0, column=0, pady=5)
        self.serveInLabel.grid(row=0, column=1, pady=5)
        self.serveOutLabel.grid(row=0, column=2, pady=5)
        self.serveShortLabel.grid(row=0, column=3, pady=5)
        
        self.receiveTargetedLabel.grid(row=0, column=0, pady=5)
        self.receiveHighLabel.grid(row=0, column=1, pady=5)
        self.receiveOffLabel.grid(row=0, column=2, pady=5)
        self.receiveLowLabel.grid(row=0, column=3, pady=5)
        
        self.setTargetedLabel.grid(row=0, column=0, pady=5)
        self.setHighLabel.grid(row=0, column=1, pady=5)
        self.setOffLabel.grid(row=0, column=2, pady=5)
        self.setLowLabel.grid(row=0, column=3, pady=5)
        
        self.spikeSuccessLabel.grid(row=0, column=0, pady=5)
        self.spikeInLabel.grid(row=0, column=1, pady=5)
        self.spikeBlockedLabel.grid(row=0, column=2, pady=5)
        self.spikeOutLabel.grid(row=0, column=3, pady=5)
        
        self.tipSuccessLabel.grid(row=0, column=0, pady=(5, 15))
        self.tipInLabel.grid(row=0, column=1, pady=(5, 15))
        self.tipBlockedLabel.grid(row=0, column=2, pady=(5, 15))
        self.tipOutLabel.grid(row=0, column=3, pady=(5, 15))
        
        self.blockScoreLabel.grid(row=0, column=0, pady=(5, 15))
        self.blockTouchLabel.grid(row=0, column=1, pady=(5, 15))
        self.blockOffLabel.grid(row=0, column=2, pady=(5, 15))
        self.blockFailLabel.grid(row=0, column=3, pady=(5, 15))
        
        self.FaultsLabel.grid(row=1, column=0, columnspan=2, pady=4)
        

        self.wb.save('data/volley_stats.xlsx')

        return
    
    def addPlayerWindow(self):
        top = Toplevel()
        top.title("Add New Player")
        nameLabel = Label(top, text="Insert Player's Full Name:")
        nameField = Entry(top, width=20)
        nicknameLabel = Label(top, text="Player Nickname (For Selection Button):")
        nicknameField = Entry(top, width=20)
        sureButton = Button(top, text= "OK", command=lambda:self.addPlayer(nameField.get(), nicknameField.get(), top), padx=20, pady=10)
        cancelButton = Button(top, text= "Cancel", command=lambda:top.destroy(), padx=20, pady=10)
        
        nameLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        nameField.grid(row=1, column=0, columnspan=2)
        nicknameLabel.grid(row=2, column=0, columnspan=2, padx=20, pady=(20, 10))
        nicknameField.grid(row=3, column=0, columnspan=2)
        sureButton.grid(row=4, column=0, padx=10, pady=10)
        cancelButton.grid(row=4, column=1, padx=10, pady=10)
        return

    def addPlayer(self, playerName, playerNickname, top):
        if(playerName is None or playerNickname is None):
            messagebox.showinfo("Error", "Both Fields must be filled")
            return
        if(not ("".join(playerName.split())).isalpha() and not ("".join(playerNickname.split())).isalpha()):
            messagebox.showinfo("Error", "Name can only contain standard characters")
            return

        self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value = self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value + 1
        self.playerNumber = self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value

        self.playerList.append(playerName)
        self.playerList.sort()
        for i in range(self.playerList.index(playerName), len(self.playerList)-1):
            self.buttonList[i].grid_forget()
        self.playerNicknameList.insert(self.playerList.index(playerName), playerNickname)

        self.buttonList.insert(self.playerList.index(playerName), Button(self.playerSelection, text=self.playerNicknameList[self.playerList.index(playerName)], command=lambda x=self.playerList.index(playerName): self.playerSelect(x), height=4, width=15))
        self.buttonReset()

        

        colCount = 0
        for col in self.ws2.iter_cols(None, None, ((self.seasonNumber-1) * 7) + 6, ((self.seasonNumber-1) * 7) + 6):
            for cell in col:
                cell.value = self.playerList[colCount]
                cell.alignment = Alignment(wrap_text=True)
            colCount = colCount + 1
            if(colCount >= len(self.playerList)):
                break
        
        colCount = 0
        for col in self.ws2.iter_cols(None, None, ((self.seasonNumber-1) * 7) + 8, ((self.seasonNumber-1) * 7) + 8):
            for cell in col:
                cell.value = self.playerNicknameList[colCount]
                cell.alignment = Alignment(wrap_text=True)
            colCount = colCount + 1
            if(colCount >= len(self.playerList)):
                break

        for i in range(len(self.weekList)):
            self.ws1.insert_rows(3 + self.playerList.index(playerName) + (i * (3 + len(self.playerList))))
            self.ws1['A' + str(3 + self.playerList.index(playerName) + (i * (3 + len(self.playerList))))].value = playerName
            for cell in self.ws1['B'+ str(3 + self.playerList.index(playerName) + (i * (3 + len(self.playerList)))):'AL' + str(3 + self.playerList.index(playerName) + (i * (3 + len(self.playerList))))]:
                for k in cell:
                    k.value = 0
        
        self.wb.save('data/volley_stats.xlsx')
        top.destroy()
        return

    def removePlayerWindow(self):
        top = Toplevel()
        top.title("Remove Player")
        nameLabel = Label(top, text="Insert Player's Full Name")
        nameField = Entry(top, width=20)
        sureButton = Button(top, text= "OK", command=lambda:self.removePlayer(nameField.get(), top), padx=20, pady=10)
        cancelButton = Button(top, text= "Cancel", command=lambda:top.destroy(), padx=20, pady=10)
        
        nameLabel.grid(row=0, column=0, columnspan=2, padx=20, pady=(20, 10))
        nameField.grid(row=1, column=0, columnspan=2)
        sureButton.grid(row=2, column=0, padx=10, pady=10)
        cancelButton.grid(row=2, column=1, padx=10, pady=10)
        return

    def removePlayer(self, playerName, top):
        if(not (playerName in self.playerList)):
            messagebox.showinfo("Error", "Player not found")
            return

        self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value = self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value - 1
        self.playerNumber = self.ws2[('B' + str(((self.seasonNumber-1) * 7) + 4))].value
        
        for i in range(len(self.weekList)):
            self.ws1.delete_rows(3 + self.playerList.index(playerName) + (i * (2 + len(self.playerList))))  

        self.playerNicknameList.pop(self.playerList.index(playerName))
        for i in range(len(self.buttonList)):
            self.buttonList[i].grid_forget()
        self.buttonList.pop(self.playerList.index(playerName))
        self.playerList.remove(playerName)
        self.buttonReset()

        self.ws2.delete_rows(((self.seasonNumber-1) * 7) + 6)
        self.ws2.insert_rows(((self.seasonNumber-1) * 7) + 6)
        colCount = 0
        for col in self.ws2.iter_cols(None, None, ((self.seasonNumber-1) * 7) + 6, ((self.seasonNumber-1) * 7) + 6):
            if(colCount >= len(self.playerList)):
                break
            for cell in col:
                cell.value = self.playerList[colCount]
            colCount = colCount + 1

        self.ws2.delete_rows(((self.seasonNumber-1) * 7) + 8)
        self.ws2.insert_rows(((self.seasonNumber-1) * 7) + 8)
        colCount = 0
        for col in self.ws2.iter_cols(None, None, ((self.seasonNumber-1) * 7) + 8, ((self.seasonNumber-1) * 7) + 8):
            if(colCount >= len(self.playerList)):
                break
            for cell in col:
                cell.value = self.playerNicknameList[colCount]
            colCount = colCount + 1
  
        self.wb.save('data/volley_stats.xlsx')
        messagebox.showinfo("Success", "Player Successfully Removed")
        top.destroy()
        return

    def toggleStatButtons(self, master):
        if self.toggleBool == False:
            self.serveAceButton = Button(self.servesFrame, text="Ace", command=lambda: self.statDecrease("serveAce"), height=3, width=10)
            self.serveInButton = Button(self.servesFrame, text="In", command=lambda: self.statDecrease("serveIn"), height=3, width=10)
            self.serveOutButton = Button(self.servesFrame, text="Out", command=lambda: self.statDecrease("serveOut"), height=3, width=10)
            self.serveShortButton = Button(self.servesFrame, text="Short", command=lambda: self.statDecrease("serveShort"), height=3, width=10)
        

            self.receiveTargetedButton = Button(self.receivesFrame, text="Targeted", command=lambda: self.statDecrease("receiveTargeted"), height=3, width=10)
            self.receiveHighButton = Button(self.receivesFrame, text="High", command=lambda: self.statDecrease("receiveHigh"), height=3, width=10)
            self.receiveOffButton = Button(self.receivesFrame, text="Off", command=lambda: self.statDecrease("receiveOff"), height=3, width=10)
            self.receiveLowButton = Button(self.receivesFrame, text="Low", command=lambda: self.statDecrease("receiveLow"), height=3, width=10)


            self.setTargetedButton = Button(self.setsFrame, text="Targeted", command=lambda: self.statDecrease("setTargeted"), height=3, width=10)
            self.setHighButton = Button(self.setsFrame, text="High", command=lambda: self.statDecrease("setHigh"), height=3, width=10)
            self.setOffButton = Button(self.setsFrame, text="Off", command=lambda: self.statDecrease("setOff"), height=3, width=10)
            self.setLowButton = Button(self.setsFrame, text="Low", command=lambda: self.statDecrease("setLow"), height=3, width=10)


            self.spikeSuccessButton = Button(self.spikesFrame, text="Success", command=lambda: self.statDecrease("spikeSuccess"), height=3, width=10)
            self.spikeInButton = Button(self.spikesFrame, text="In", command=lambda: self.statDecrease("spikeIn"), height=3, width=10)
            self.spikeBlockedButton = Button(self.spikesFrame, text="Blocked", command=lambda: self.statDecrease("spikeBlocked"), height=3, width=10)
            self.spikeOutButton = Button(self.spikesFrame, text="Out", command=lambda: self.statDecrease("spikeOut"), height=3, width=10)
            

            self.tipSuccessButton = Button(self.tipsFrame, text="Success", command=lambda: self.statDecrease("tipSuccess"), height=3, width=10)
            self.tipInButton = Button(self.tipsFrame, text="In", command=lambda: self.statDecrease("tipIn"), height=3, width=10)
            self.tipBlockedButton = Button(self.tipsFrame, text="Blocked", command=lambda: self.statDecrease("tipBlocked"), height=3, width=10)
            self.tipOutButton = Button(self.tipsFrame, text="Out", command=lambda: self.statDecrease("tipOut"), height=3, width=10)


            self.blockScoreButton = Button(self.blocksFrame, text="Score", command=lambda: self.statDecrease("blockScore"), height=3, width=10)
            self.blockTouchButton = Button(self.blocksFrame, text="Touch", command=lambda: self.statDecrease("blockTouch"), height=3, width=10)
            self.blockOffButton = Button(self.blocksFrame, text="Off", command=lambda: self.statDecrease("blockOff"), height=3, width=10)
            self.blockFailButton = Button(self.blocksFrame, text="No Block", command=lambda: self.statDecrease("blockFail"), height=3, width=10)

            self.toggleButton = Button(master,  text="Toggle \n Increase", command=lambda:self.toggleStatButtons(master), bg='Red', height=3, width=10)

        if self.toggleBool == True:
            self.serveAceButton = Button(self.servesFrame, text="Ace", command=lambda: self.statIncrease("serveAce"), height=3, width=10)
            self.serveInButton = Button(self.servesFrame, text="In", command=lambda: self.statIncrease("serveIn"), height=3, width=10)
            self.serveOutButton = Button(self.servesFrame, text="Out", command=lambda: self.statIncrease("serveOut"), height=3, width=10)
            self.serveShortButton = Button(self.servesFrame, text="Short", command=lambda: self.statIncrease("serveShort"), height=3, width=10)
        

            self.receiveTargetedButton = Button(self.receivesFrame, text="Targeted", command=lambda: self.statIncrease("receiveTargeted"), height=3, width=10)
            self.receiveHighButton = Button(self.receivesFrame, text="High", command=lambda: self.statIncrease("receiveHigh"), height=3, width=10)
            self.receiveOffButton = Button(self.receivesFrame, text="Off", command=lambda: self.statIncrease("receiveOff"), height=3, width=10)
            self.receiveLowButton = Button(self.receivesFrame, text="Low", command=lambda: self.statIncrease("receiveLow"), height=3, width=10)


            self.setTargetedButton = Button(self.setsFrame, text="Targeted", command=lambda: self.statIncrease("setTargeted"), height=3, width=10)
            self.setHighButton = Button(self.setsFrame, text="High", command=lambda: self.statIncrease("setHigh"), height=3, width=10)
            self.setOffButton = Button(self.setsFrame, text="Off", command=lambda: self.statIncrease("setOff"), height=3, width=10)
            self.setLowButton = Button(self.setsFrame, text="Low", command=lambda: self.statIncrease("setLow"), height=3, width=10)


            self.spikeSuccessButton = Button(self.spikesFrame, text="Success", command=lambda: self.statIncrease("spikeSuccess"), height=3, width=10)
            self.spikeInButton = Button(self.spikesFrame, text="In", command=lambda: self.statIncrease("spikeIn"), height=3, width=10)
            self.spikeBlockedButton = Button(self.spikesFrame, text="Blocked", command=lambda: self.statIncrease("spikeBlocked"), height=3, width=10)
            self.spikeOutButton = Button(self.spikesFrame, text="Out", command=lambda: self.statIncrease("spikeOut"), height=3, width=10)
            

            self.tipSuccessButton = Button(self.tipsFrame, text="Success", command=lambda: self.statIncrease("tipSuccess"), height=3, width=10)
            self.tipInButton = Button(self.tipsFrame, text="In", command=lambda: self.statIncrease("tipIn"), height=3, width=10)
            self.tipBlockedButton = Button(self.tipsFrame, text="Blocked", command=lambda: self.statIncrease("tipBlocked"), height=3, width=10)
            self.tipOutButton = Button(self.tipsFrame, text="Out", command=lambda: self.statIncrease("tipOut"), height=3, width=10)


            self.blockScoreButton = Button(self.blocksFrame, text="Score", command=lambda: self.statIncrease("blockScore"), height=3, width=10)
            self.blockTouchButton = Button(self.blocksFrame, text="Touch", command=lambda: self.statIncrease("blockTouch"), height=3, width=10)
            self.blockOffButton = Button(self.blocksFrame, text="Off", command=lambda: self.statIncrease("blockOff"), height=3, width=10)
            self.blockFailButton = Button(self.blocksFrame, text="No Block", command=lambda: self.statIncrease("blockFail"), height=3, width=10)

            self.toggleButton = Button(master, text="Toggle \n Decrease", command=lambda:self.toggleStatButtons(master), height=3, width=10)
        
        self.serveAceButton.grid_forget()
        self.serveInButton.grid_forget()
        self.serveOutButton.grid_forget()
        self.serveShortButton.grid_forget()

        
        self.receiveTargetedButton.grid_forget()
        self.receiveHighButton.grid_forget()
        self.receiveOffButton.grid_forget()
        self.receiveLowButton.grid_forget()

        
        self.setTargetedButton.grid_forget()
        self.setHighButton.grid_forget()
        self.setOffButton.grid_forget()
        self.setLowButton.grid_forget()

        
        self.spikeSuccessButton.grid_forget()
        self.spikeInButton.grid_forget()
        self.spikeBlockedButton.grid_forget()
        self.spikeOutButton.grid_forget()


        self.tipSuccessButton.grid_forget()
        self.tipInButton.grid_forget()
        self.tipBlockedButton.grid_forget()
        self.tipOutButton.grid_forget()

        
        self.blockScoreButton.grid_forget()
        self.blockTouchButton.grid_forget()
        self.blockOffButton.grid_forget()
        self.blockFailButton.grid_forget()

        self.toggleButton.grid_forget()

        self.serveAceButton.grid(row=0, column=0)
        self.serveInButton.grid(row=0, column=1)
        self.serveOutButton.grid(row=0, column=2)
        self.serveShortButton.grid(row=0, column=3)

        
        self.receiveTargetedButton.grid(row=0, column=0)
        self.receiveHighButton.grid(row=0, column=1)
        self.receiveOffButton.grid(row=0, column=2)
        self.receiveLowButton.grid(row=0, column=3)

        
        self.setTargetedButton.grid(row=0, column=0)
        self.setHighButton.grid(row=0, column=1)
        self.setOffButton.grid(row=0, column=2)
        self.setLowButton.grid(row=0, column=3)

        
        self.spikeSuccessButton.grid(row=0, column=0)
        self.spikeInButton.grid(row=0, column=1)
        self.spikeBlockedButton.grid(row=0, column=2)
        self.spikeOutButton.grid(row=0, column=3)


        self.tipSuccessButton.grid(row=0, column=0)
        self.tipInButton.grid(row=0, column=1)
        self.tipBlockedButton.grid(row=0, column=2)
        self.tipOutButton.grid(row=0, column=3)

        
        self.blockScoreButton.grid(row=0, column=0)
        self.blockTouchButton.grid(row=0, column=1)
        self.blockOffButton.grid(row=0, column=2)
        self.blockFailButton.grid(row=0, column=3)

        self.toggleButton.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

        self.toggleBool = not self.toggleBool
        return